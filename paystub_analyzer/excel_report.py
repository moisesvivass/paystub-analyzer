import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from paystub_analyzer.config import OUTPUT_EXCEL
from paystub_analyzer.logger import get_logger

logger = get_logger(__name__)

# ── Palette ────────────────────────────────────────────────────────────────────
DARK_BLUE   = "1F3864"
MID_BLUE    = "2E75B6"
LIGHT_BLUE  = "D6E4F0"
LIGHT_GRAY  = "F2F2F2"
WHITE       = "FFFFFF"
GREEN       = "1E8449"
LIGHT_GREEN = "D5F5E3"

HEADERS = [
    "Company", "Pay Period Start", "Pay Period End", "Gross Pay",
    "Net Pay", "Federal Tax", "Provincial Tax", "CPP", "EI",
    "Vacation Pay", "Hours Worked",
]

# ── Style helpers ──────────────────────────────────────────────────────────────
def _thin_border() -> Border:
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _title_cell(ws, ref: str, text: str, size: int = 13) -> None:
    c = ws[ref]
    c.value = text
    c.font = Font(name="Arial", bold=True, size=size, color=WHITE)
    c.fill = PatternFill("solid", fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")


def _header_cell(cell, text: str) -> None:
    cell.value = text
    cell.font = Font(name="Arial", bold=True, size=10, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=MID_BLUE)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _thin_border()


def _data_cell(cell, value: object, fmt: str | None = None,
               bold: bool = False, bg: str | None = None) -> None:
    cell.value = value
    cell.font = Font(name="Arial", size=10, bold=bold)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _thin_border()
    if fmt:
        cell.number_format = fmt
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)


def _set_col_widths(ws, widths: dict[str, float]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ── Load existing data ─────────────────────────────────────────────────────────
def load_existing_data() -> list[list]:
    if not os.path.exists(OUTPUT_EXCEL):
        return []
    wb = openpyxl.load_workbook(OUTPUT_EXCEL, read_only=True, data_only=True)
    sheet_name = "📋 Raw Data" if "📋 Raw Data" in wb.sheetnames else None
    if sheet_name is None:
        ws = wb.active
        min_row = 2
    else:
        ws = wb[sheet_name]
        min_row = 4
    existing = [list(row) for row in ws.iter_rows(min_row=min_row, values_only=True) if any(row)]
    wb.close()
    logger.info(f"Loaded {len(existing)} existing records")
    return existing


# ── Deduplicate ────────────────────────────────────────────────────────────────
def deduplicate(existing: list[list], new_data: list[dict]) -> list[list]:
    """Merge new paystubs into existing rows, deduplicating by (company, pay_period_end)."""
    existing_keys: set[tuple] = {(row[0], row[2]) for row in existing if len(row) >= 3}
    rows = [list(row) for row in existing]

    added = 0
    for d in new_data:
        key = (d.get("company"), d.get("pay_period_end"))
        if key not in existing_keys:
            existing_keys.add(key)
            rows.append([
                d.get("company"), d.get("pay_period_start"), d.get("pay_period_end"),
                d.get("gross_pay"), d.get("net_pay"), d.get("federal_tax"),
                d.get("provincial_tax"), d.get("cpp"), d.get("ei"),
                d.get("vacation_pay"), d.get("hours_worked"),
            ])
            added += 1

    rows.sort(key=lambda x: x[1] or "")
    logger.info(f"{added} new records added — {len(rows)} total")
    return rows


# ── Sheet builders ─────────────────────────────────────────────────────────────
def _build_raw_data(wb: openpyxl.Workbook, rows: list[list]) -> None:
    ws = wb.create_sheet("📋 Raw Data")
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A1:K1")
    _title_cell(ws, "A1", "📋  RAW DATA — ALL PAY PERIODS")

    for col, h in enumerate(HEADERS, 1):
        _header_cell(ws.cell(row=3, column=col), h)

    for r, row in enumerate(rows, 4):
        bg = LIGHT_GRAY if r % 2 == 0 else WHITE
        for c, val in enumerate(row, 1):
            fmt = "$#,##0.00" if c in (4, 5, 6, 7, 8, 9, 10) else None
            _data_cell(ws.cell(row=r, column=c), val, fmt=fmt, bg=bg)

    _set_col_widths(ws, {"A": 36, "B": 18, "C": 18, "D": 14, "E": 12,
                         "F": 14, "G": 16, "H": 10, "I": 10, "J": 14, "K": 14})
    ws.freeze_panes = "A4"


def _build_annual_summary(wb: openpyxl.Workbook, rows: list[list]) -> None:
    ws = wb.create_sheet("📊 Annual Summary")
    ws.merge_cells("A1:H1")
    ws.row_dimensions[1].height = 30
    _title_cell(ws, "A1", "📊  ANNUAL SUMMARY — EARNINGS BY YEAR")

    for c, h in enumerate(["Year", "Pay Periods", "Gross Pay", "Net Pay",
                            "Federal Tax", "Provincial Tax", "CPP", "EI"], 1):
        _header_cell(ws.cell(row=3, column=c), h)

    years: dict[str, list] = {}
    for row in rows:
        y = str(row[1])[:4] if row[1] else "Unknown"
        if y not in years:
            years[y] = [0] * 7
        years[y][0] += 1
        for i, idx in enumerate([3, 4, 5, 6, 7, 8]):
            try:
                years[y][i + 1] += float(row[idx] or 0)
            except (TypeError, ValueError):
                pass

    for r, (y, vals) in enumerate(sorted(years.items()), 4):
        bg = LIGHT_BLUE if r % 2 == 0 else WHITE
        _data_cell(ws.cell(row=r, column=1), y, bg=bg)
        _data_cell(ws.cell(row=r, column=2), vals[0], bg=bg)
        for c, v in enumerate(vals[1:], 3):
            _data_cell(ws.cell(row=r, column=c), v, fmt="$#,##0.00", bg=bg)

    tr = 4 + len(years)
    _data_cell(ws.cell(row=tr, column=1), "TOTAL", bold=True, bg=LIGHT_GREEN)
    _data_cell(ws.cell(row=tr, column=2), f"=SUM(B4:B{tr-1})", bold=True, bg=LIGHT_GREEN)
    for c in range(3, 9):
        col = get_column_letter(c)
        _data_cell(ws.cell(row=tr, column=c),
                   f"=SUM({col}4:{col}{tr-1})", fmt="$#,##0.00", bold=True, bg=LIGHT_GREEN)

    _set_col_widths(ws, {"A": 12, "B": 14, "C": 14, "D": 14,
                         "E": 14, "F": 16, "G": 12, "H": 12})


def _build_monthly_summary(wb: openpyxl.Workbook, rows: list[list]) -> None:
    ws = wb.create_sheet("📅 Monthly Summary")
    ws.merge_cells("A1:G1")
    ws.row_dimensions[1].height = 30
    _title_cell(ws, "A1", "📅  MONTHLY SUMMARY — EARNINGS OVER TIME")

    for c, h in enumerate(["Year", "Month", "Pay Periods", "Gross Pay",
                            "Net Pay", "Total Tax", "Net Rate %"], 1):
        _header_cell(ws.cell(row=3, column=c), h)

    _MONTHS = {"01": "Jan", "02": "Feb", "03": "Mar", "04": "Apr",
               "05": "May", "06": "Jun", "07": "Jul", "08": "Aug",
               "09": "Sep", "10": "Oct", "11": "Nov", "12": "Dec"}

    monthly: dict[str, dict] = {}
    for row in rows:
        if not row[1]:
            continue
        key = str(row[1])[:7]
        y, m = key[:4], key[5:7]
        if key not in monthly:
            monthly[key] = {"year": y, "month": _MONTHS.get(m, m),
                            "periods": 0, "gross": 0.0, "net": 0.0, "tax": 0.0}
        monthly[key]["periods"] += 1
        try:
            monthly[key]["gross"] += float(row[3] or 0)
            monthly[key]["net"] += float(row[4] or 0)
            monthly[key]["tax"] += float(row[5] or 0) + float(row[6] or 0)
        except (TypeError, ValueError):
            pass

    for r, (_, v) in enumerate(sorted(monthly.items()), 4):
        bg = LIGHT_GRAY if r % 2 == 0 else WHITE
        _data_cell(ws.cell(row=r, column=1), v["year"], bg=bg)
        _data_cell(ws.cell(row=r, column=2), v["month"], bg=bg)
        _data_cell(ws.cell(row=r, column=3), v["periods"], bg=bg)
        _data_cell(ws.cell(row=r, column=4), v["gross"], fmt="$#,##0.00", bg=bg)
        _data_cell(ws.cell(row=r, column=5), v["net"], fmt="$#,##0.00", bg=bg)
        _data_cell(ws.cell(row=r, column=6), v["tax"], fmt="$#,##0.00", bg=bg)
        rate = (v["net"] / v["gross"] * 100) if v["gross"] else 0
        _data_cell(ws.cell(row=r, column=7), round(rate, 1), fmt="0.0%", bg=bg)

    _set_col_widths(ws, {"A": 10, "B": 10, "C": 13, "D": 14,
                         "E": 14, "F": 14, "G": 12})


def _build_by_company(wb: openpyxl.Workbook, rows: list[list]) -> None:
    ws = wb.create_sheet("🏢 By Company")
    ws.merge_cells("A1:G1")
    ws.row_dimensions[1].height = 30
    _title_cell(ws, "A1", "🏢  EARNINGS BY COMPANY")

    for c, h in enumerate(["Company", "Pay Periods", "Gross Pay", "Net Pay",
                            "Avg Gross/Period", "Avg Net/Period", "Avg Hours"], 1):
        _header_cell(ws.cell(row=3, column=c), h)

    companies: dict[str, dict] = {}
    for row in rows:
        co = row[0] or "Unknown"
        if co not in companies:
            companies[co] = {"periods": 0, "gross": 0.0, "net": 0.0, "hours": 0.0}
        companies[co]["periods"] += 1
        try:
            companies[co]["gross"] += float(row[3] or 0)
            companies[co]["net"] += float(row[4] or 0)
            companies[co]["hours"] += float(row[10] or 0)
        except (TypeError, ValueError):
            pass

    for r, (co, v) in enumerate(sorted(companies.items()), 4):
        bg = LIGHT_BLUE if r % 2 == 0 else WHITE
        p = v["periods"]
        _data_cell(ws.cell(row=r, column=1), co, bg=bg)
        _data_cell(ws.cell(row=r, column=2), p, bg=bg)
        _data_cell(ws.cell(row=r, column=3), v["gross"], fmt="$#,##0.00", bg=bg)
        _data_cell(ws.cell(row=r, column=4), v["net"], fmt="$#,##0.00", bg=bg)
        _data_cell(ws.cell(row=r, column=5), round(v["gross"] / p, 2) if p else 0, fmt="$#,##0.00", bg=bg)
        _data_cell(ws.cell(row=r, column=6), round(v["net"] / p, 2) if p else 0, fmt="$#,##0.00", bg=bg)
        _data_cell(ws.cell(row=r, column=7), round(v["hours"] / p, 1) if p else 0, bg=bg)

    _set_col_widths(ws, {"A": 38, "B": 13, "C": 14, "D": 14,
                         "E": 18, "F": 16, "G": 13})


def _build_deductions(wb: openpyxl.Workbook, rows: list[list]) -> None:
    ws = wb.create_sheet("💰 Deductions")
    ws.merge_cells("A1:G1")
    ws.row_dimensions[1].height = 30
    _title_cell(ws, "A1", "💰  DEDUCTIONS BREAKDOWN")

    for c, h in enumerate(["Pay Period End", "Company", "Gross Pay",
                            "Federal Tax", "Provincial Tax", "CPP", "EI"], 1):
        _header_cell(ws.cell(row=3, column=c), h)

    for r, row in enumerate(rows, 4):
        bg = LIGHT_GRAY if r % 2 == 0 else WHITE
        _data_cell(ws.cell(row=r, column=1), row[2], bg=bg)
        _data_cell(ws.cell(row=r, column=2), row[0], bg=bg)
        for c, idx in enumerate([3, 5, 6, 7, 8], 3):
            _data_cell(ws.cell(row=r, column=c), row[idx], fmt="$#,##0.00", bg=bg)

    tr = 4 + len(rows)
    _data_cell(ws.cell(row=tr, column=1), "TOTAL", bold=True, bg=LIGHT_GREEN)
    _data_cell(ws.cell(row=tr, column=2), "", bg=LIGHT_GREEN)
    for c in range(3, 8):
        col = get_column_letter(c)
        _data_cell(ws.cell(row=tr, column=c),
                   f"=SUM({col}4:{col}{tr-1})", fmt="$#,##0.00", bold=True, bg=LIGHT_GREEN)

    _set_col_widths(ws, {"A": 16, "B": 36, "C": 14, "D": 14,
                         "E": 16, "F": 10, "G": 10})
    ws.freeze_panes = "A4"


def _build_dashboard(wb: openpyxl.Workbook, rows: list[list]) -> None:
    ws = wb.create_sheet("🏠 Dashboard")
    ws.merge_cells("A1:F1")
    ws.row_dimensions[1].height = 40
    _title_cell(ws, "A1", "🏠  PAYSTUB ANALYZER — EARNINGS DASHBOARD", size=14)

    total_gross = sum(float(r[3] or 0) for r in rows)
    total_net = sum(float(r[4] or 0) for r in rows)
    total_tax = sum(float(r[5] or 0) + float(r[6] or 0) for r in rows)
    total_cpp = sum(float(r[7] or 0) for r in rows)
    total_ei = sum(float(r[8] or 0) for r in rows)
    periods = len(rows)

    metrics: list[tuple] = [
        ("Total Pay Periods",  periods,                          None),
        ("Total Gross Pay",    total_gross,                      "$#,##0.00"),
        ("Total Net Pay",      total_net,                        "$#,##0.00"),
        ("Total Income Tax",   total_tax,                        "$#,##0.00"),
        ("Total CPP",          total_cpp,                        "$#,##0.00"),
        ("Total EI",           total_ei,                         "$#,##0.00"),
        ("Avg Gross / Period", total_gross / periods if periods else 0, "$#,##0.00"),
        ("Avg Net / Period",   total_net / periods if periods else 0,   "$#,##0.00"),
    ]

    ws.cell(row=3, column=1).value = "METRIC"
    ws.cell(row=3, column=1).font = Font(name="Arial", bold=True, size=11)
    ws.cell(row=3, column=2).value = "VALUE"
    ws.cell(row=3, column=2).font = Font(name="Arial", bold=True, size=11)

    for r, (label, val, fmt) in enumerate(metrics, 4):
        bg = LIGHT_BLUE if r % 2 == 0 else WHITE
        lc = ws.cell(row=r, column=1)
        vc = ws.cell(row=r, column=2)
        lc.value = label
        lc.font = Font(name="Arial", size=11)
        lc.fill = PatternFill("solid", fgColor=bg)
        lc.border = _thin_border()
        vc.value = val
        vc.font = Font(name="Arial", size=11, bold=True, color=GREEN)
        vc.fill = PatternFill("solid", fgColor=bg)
        vc.border = _thin_border()
        if fmt:
            vc.number_format = fmt

    _set_col_widths(ws, {"A": 25, "B": 18})


def _build_glossary(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("📖 Glossary")
    ws.merge_cells("A1:D1")
    ws.row_dimensions[1].height = 30
    _title_cell(ws, "A1", "📖  GLOSSARY — CANADIAN PAYSTUB TERMS")

    for c, h in enumerate(["Term", "Full Name", "Description", "Who Pays"], 1):
        _header_cell(ws.cell(row=3, column=c), h)

    terms = [
        ("Gross Pay",    "Gross Earnings",                "Total earnings before any deductions",         "N/A"),
        ("Net Pay",      "Net Earnings / Take-Home Pay",  "Amount received after all deductions",         "N/A"),
        ("Federal Tax",  "Federal Income Tax",            "Tax paid to the Government of Canada",         "Employee"),
        ("Provincial",   "Provincial Income Tax",         "Tax paid to the provincial government",        "Employee"),
        ("CPP",          "Canada Pension Plan",           "Retirement savings contribution",               "Employee & Employer"),
        ("EI",           "Employment Insurance",          "Insurance for job loss or leave",              "Employee & Employer"),
        ("Vacation Pay", "Vacation Pay Accrual",          "Usually 4% of gross, paid out or accrued",     "Employer"),
        ("Hours Worked", "Regular Hours",                 "Total hours worked in the pay period",         "N/A"),
    ]

    for r, (term, full, desc, who) in enumerate(terms, 4):
        bg = LIGHT_GRAY if r % 2 == 0 else WHITE
        for c, val in enumerate([term, full, desc, who], 1):
            cell = ws.cell(row=r, column=c)
            cell.value = val
            cell.font = Font(name="Arial", size=10, bold=(c == 1))
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.border = _thin_border()
            cell.alignment = Alignment(wrap_text=True, vertical="center")

    _set_col_widths(ws, {"A": 16, "B": 30, "C": 50, "D": 22})


# ── Year Personal Summary (reusable) ──────────────────────────────────────────
def _build_year_personal(wb: openpyxl.Workbook, rows: list[list], year: str, icon: str) -> None:
    year_rows = [r for r in rows if str(r[1] or "").startswith(year)]

    ws = wb.create_sheet(f"{icon} {year} Personal")
    ws.merge_cells("A1:G1")
    ws.row_dimensions[1].height = 40
    _title_cell(ws, "A1", f"{icon}  {year} — PERSONAL EARNINGS SUMMARY", size=14)

    if not year_rows:
        ws["A3"].value = f"No paystubs found for {year} yet."
        ws["A3"].font = Font(name="Arial", size=11, italic=True)
        return

    total_gross = sum(float(r[3] or 0) for r in year_rows)
    total_net   = sum(float(r[4] or 0) for r in year_rows)
    total_tax   = sum(float(r[5] or 0) + float(r[6] or 0) for r in year_rows)
    total_cpp   = sum(float(r[7] or 0) for r in year_rows)
    total_ei    = sum(float(r[8] or 0) for r in year_rows)
    periods     = len(year_rows)
    ytd_rate    = round((total_net / total_gross * 100), 1) if total_gross else 0

    summary = [
        ("Pay Periods",      periods,                                   None),
        ("Gross Pay",        total_gross,                               "$#,##0.00"),
        ("Net Pay",          total_net,                                 "$#,##0.00"),
        ("Income Tax",       total_tax,                                 "$#,##0.00"),
        ("CPP",              total_cpp,                                 "$#,##0.00"),
        ("EI",               total_ei,                                  "$#,##0.00"),
        ("Avg Net / Period", total_net / periods if periods else 0,     "$#,##0.00"),
        ("Net Rate",         ytd_rate,                                  "0.0\"%\""),
    ]

    ws.cell(row=3, column=1).value = "METRIC"
    ws.cell(row=3, column=1).font = Font(name="Arial", bold=True, size=11)
    ws.cell(row=3, column=2).value = "VALUE"
    ws.cell(row=3, column=2).font = Font(name="Arial", bold=True, size=11)

    for r, (label, val, fmt) in enumerate(summary, 4):
        bg = LIGHT_BLUE if r % 2 == 0 else WHITE
        lc = ws.cell(row=r, column=1)
        vc = ws.cell(row=r, column=2)
        lc.value = label
        lc.font = Font(name="Arial", size=11)
        lc.fill = PatternFill("solid", fgColor=bg)
        lc.border = _thin_border()
        vc.value = val
        vc.font = Font(name="Arial", size=11, bold=True, color=GREEN)
        vc.fill = PatternFill("solid", fgColor=bg)
        vc.border = _thin_border()
        if fmt:
            vc.number_format = fmt

    detail_row = 4 + len(summary) + 2
    ws.merge_cells(f"A{detail_row}:I{detail_row}")
    _title_cell(ws, f"A{detail_row}", "PAYSTUB DETAIL", size=11)

    for c, h in enumerate(["Pay Period End", "Gross Pay", "Net Pay", "Federal Tax",
                            "Provincial Tax", "CPP", "EI", "Vacation Pay", "Hours Worked"], 1):
        _header_cell(ws.cell(row=detail_row + 1, column=c), h)

    for i, row in enumerate(sorted(year_rows, key=lambda x: x[2] or ""), detail_row + 2):
        bg = LIGHT_GRAY if i % 2 == 0 else WHITE
        _data_cell(ws.cell(row=i, column=1), row[2], bg=bg)
        for c, idx in enumerate([3, 4, 5, 6, 7, 8], 2):
            _data_cell(ws.cell(row=i, column=c), row[idx], fmt="$#,##0.00", bg=bg)
        _data_cell(ws.cell(row=i, column=8), row[9], fmt="$#,##0.00", bg=bg)
        _data_cell(ws.cell(row=i, column=9), row[10], bg=bg)

    tr = detail_row + 2 + len(year_rows)
    _data_cell(ws.cell(row=tr, column=1), "TOTAL", bold=True, bg=LIGHT_GREEN)
    for c in range(2, 9):
        col = get_column_letter(c)
        _data_cell(ws.cell(row=tr, column=c),
                   f"=SUM({col}{detail_row + 2}:{col}{tr - 1})",
                   fmt="$#,##0.00", bold=True, bg=LIGHT_GREEN)
    _data_cell(ws.cell(row=tr, column=9), "", bg=LIGHT_GREEN)

    _set_col_widths(ws, {"A": 18, "B": 14, "C": 14, "D": 14, "E": 16, "F": 10, "G": 10, "H": 14, "I": 14})
    ws.freeze_panes = f"A{detail_row + 2}"


# ── Public API ─────────────────────────────────────────────────────────────────
def create_excel(data_list: list[dict]) -> None:
    """Build the full Excel workbook from paystub data."""
    existing = load_existing_data()
    all_rows = deduplicate(existing, data_list)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _build_year_personal(wb, all_rows, "2026", "⭐")
    _build_year_personal(wb, all_rows, "2025", "📅")
    _build_dashboard(wb, all_rows)
    _build_raw_data(wb, all_rows)
    _build_annual_summary(wb, all_rows)
    _build_monthly_summary(wb, all_rows)
    _build_by_company(wb, all_rows)
    _build_deductions(wb, all_rows)
    _build_glossary(wb)

    wb.save(OUTPUT_EXCEL)
    logger.info(f"Excel saved: {OUTPUT_EXCEL}")
