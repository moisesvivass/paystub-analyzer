import os
import pytest
import openpyxl
from paystub_analyzer.excel_report import deduplicate, create_excel


SAMPLE_ROWS = [
    ["Acme Corp", "2025-01-01", "2025-01-15", 2000.0, 1500.0, 250.0, 150.0, 80.0, 20.0, 80.0, 80.0],
    ["Acme Corp", "2025-01-16", "2025-01-31", 2100.0, 1580.0, 260.0, 160.0, 83.0, 21.0, 84.0, 80.0],
]

SAMPLE_NEW = [
    {
        "company": "Acme Corp",
        "pay_period_start": "2025-02-01",
        "pay_period_end": "2025-02-15",
        "gross_pay": 2200.0,
        "net_pay": 1650.0,
        "federal_tax": 270.0,
        "provincial_tax": 170.0,
        "cpp": 85.0,
        "ei": 22.0,
        "vacation_pay": 88.0,
        "hours_worked": 80.0,
    }
]


def test_deduplicate_adds_new_records():
    result = deduplicate(SAMPLE_ROWS, SAMPLE_NEW)
    assert len(result) == 3


def test_deduplicate_ignores_duplicates():
    duplicate = [{
        "company": "Acme Corp",
        "pay_period_start": "2025-01-01",
        "pay_period_end": "2025-01-15",
        "gross_pay": 2000.0, "net_pay": 1500.0, "federal_tax": 0.0,
        "provincial_tax": 0.0, "cpp": 0.0, "ei": 0.0,
        "vacation_pay": 0.0, "hours_worked": None,
    }]
    result = deduplicate(SAMPLE_ROWS, duplicate)
    assert len(result) == 2


def test_deduplicate_sorts_by_pay_period_start():
    reversed_rows = list(reversed(SAMPLE_ROWS))
    result = deduplicate(reversed_rows, [])
    assert result[0][1] <= result[1][1]


def test_create_excel_produces_file(tmp_path, monkeypatch):
    monkeypatch.setenv("OUTPUT_EXCEL", str(tmp_path / "test_report.xlsx"))
    import importlib
    import paystub_analyzer.config as cfg
    importlib.reload(cfg)
    import paystub_analyzer.excel_report as er
    importlib.reload(er)

    er.create_excel(SAMPLE_NEW)

    out = tmp_path / "test_report.xlsx"
    assert out.exists()
    wb = openpyxl.load_workbook(str(out))
    assert "📋 Raw Data" in wb.sheetnames
    assert "🏠 Dashboard" in wb.sheetnames
    assert "📊 Annual Summary" in wb.sheetnames
